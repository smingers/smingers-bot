"""
Content Extractor

Multi-backend HTML content extraction with site-specific configs:
1. Site-specific CSS selectors (NYT, Reuters, Bloomberg, etc.)
2. Trafilatura (best for news articles)
3. Readability (fallback)
4. BoilerPy3 (last resort)

Also handles non-HTML formats via SpreadsheetParser:
- CSV files (detected by Content-Type or .csv extension) — stdlib csv
- Excel files (.xlsx, detected by Content-Type or extension) — openpyxl
- Google Sheets (public) — URL rewritten to CSV export before fetching
- PDF files (detected by Content-Type or .pdf extension) — pymupdf
- JSON data endpoints (detected by Content-Type or .json extension) — stdlib json

Also includes ConcurrentContentExtractor for concurrent URL fetching.
"""

import asyncio
import csv
import html as html_lib
import io
import json
import logging
import re
from dataclasses import dataclass
from typing import Any
from urllib.parse import unquote, urlparse

import httpx
from bs4 import BeautifulSoup, Tag

# Optional imports with graceful fallbacks
try:
    from trafilatura import extract as trafilatura_extract
    from trafilatura.settings import use_config

    HAS_TRAFILATURA = True
except ImportError:
    HAS_TRAFILATURA = False

try:
    from readability import Document

    HAS_READABILITY = True
except ImportError:
    HAS_READABILITY = False

try:
    from boilerpy3 import extractors

    HAS_BOILERPY = True
except ImportError:
    HAS_BOILERPY = False

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import fitz  # pymupdf

    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False

logger = logging.getLogger(__name__)


class SpreadsheetParser:
    """
    Converts tabular data (CSV text or Excel bytes) to LLM-readable markdown tables.

    Row limit prevents excessive token usage. When truncated, a note is appended
    showing how many rows were omitted.

    All parse_* methods return None on failure (library unavailable or parse error).
    Callers should treat None as a failure signal and set success=False.
    """

    MAX_ROWS = 200
    MAX_SHEETS = 10
    MAX_TOTAL_ROWS = 500

    @staticmethod
    def _to_markdown_table(headers: list[str], rows: list[list[str]]) -> str:
        """Format headers and rows as a pipe-delimited markdown table."""

        # Escape pipe characters in cell values
        def escape(cell: str) -> str:
            return str(cell).replace("|", "\\|").replace("\n", " ").strip()

        header_row = "| " + " | ".join(escape(h) for h in headers) + " |"
        separator = "| " + " | ".join("---" for _ in headers) + " |"
        data_rows = ["| " + " | ".join(escape(c) for c in row) + " |" for row in rows]
        return "\n".join([header_row, separator] + data_rows)

    @classmethod
    def parse_csv(cls, text: str, max_rows: int = MAX_ROWS) -> str | None:
        """
        Parse CSV text and return a markdown table string, or None on parse failure.

        Uses csv.Sniffer to auto-detect the delimiter (comma, semicolon, tab, pipe)
        with a fallback to standard comma-delimited if sniffing fails.

        Args:
            text: Raw CSV text content
            max_rows: Maximum number of data rows to include

        Returns:
            Markdown table string (with truncation note if needed), or None on error
        """
        try:
            # Sniff delimiter from first 4KB to handle semicolon/tab/pipe variants
            try:
                dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel  # fallback: standard comma-delimited

            reader = csv.DictReader(io.StringIO(text), dialect=dialect)
            headers = reader.fieldnames or []
            if not headers:
                return "(CSV file had no headers or was empty)"

            all_rows = []
            for row in reader:
                all_rows.append([row.get(h, "") for h in headers])

            total_rows = len(all_rows)
            truncated = total_rows > max_rows
            display_rows = all_rows[:max_rows]

            table = cls._to_markdown_table(list(headers), display_rows)
            if truncated:
                table += f"\n[Showing {max_rows} of {total_rows} rows]"
            return table
        except Exception as e:
            logger.debug(f"CSV parsing failed: {e}")
            return None

    @classmethod
    def parse_excel(
        cls,
        data: bytes,
        max_rows: int = MAX_ROWS,
        max_sheets: int = MAX_SHEETS,
        max_total_rows: int = MAX_TOTAL_ROWS,
    ) -> str | None:
        """
        Parse Excel (.xlsx) bytes and return a markdown table string, or None on parse failure.

        Reads up to max_sheets sheets; if multiple sheets exist, each is rendered as a
        separate section with a heading. Global limits (max_sheets, max_total_rows)
        prevent excessive output for large workbooks.

        Args:
            data: Raw Excel file bytes
            max_rows: Maximum data rows per sheet
            max_sheets: Maximum number of sheets to render
            max_total_rows: Maximum total data rows across all sheets

        Returns:
            Markdown string with one table per sheet (with truncation notes if needed),
            or None on parse error
        """
        if not HAS_OPENPYXL:
            logger.debug("Excel parsing unavailable: openpyxl not installed")
            return None

        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            parts = []
            total_rows_rendered = 0
            skipped_sheets = len(wb.sheetnames) - max_sheets

            for sheet_name in wb.sheetnames[:max_sheets]:
                if total_rows_rendered >= max_total_rows:
                    parts.append(f"**Sheet: {sheet_name}**\n[Skipped — global row limit reached]")
                    continue

                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)

                # First row as headers
                try:
                    raw_headers = next(rows_iter)
                except StopIteration:
                    parts.append(f"**Sheet: {sheet_name}**\n(empty sheet)")
                    continue

                headers = [str(h) if h is not None else "" for h in raw_headers]
                # Skip sheets with no meaningful header row
                if not any(h.strip() for h in headers):
                    continue

                all_rows = []
                for raw_row in rows_iter:
                    all_rows.append([str(c) if c is not None else "" for c in raw_row])

                remaining = max_total_rows - total_rows_rendered
                effective_max = min(max_rows, remaining)
                total_rows = len(all_rows)
                truncated = total_rows > effective_max
                display_rows = all_rows[:effective_max]
                total_rows_rendered += len(display_rows)

                table = cls._to_markdown_table(headers, display_rows)
                if truncated:
                    table += f"\n[Showing {effective_max} of {total_rows} rows]"

                heading = f"**Sheet: {sheet_name}**\n" if len(wb.sheetnames) > 1 else ""
                parts.append(heading + table)

            wb.close()

            if skipped_sheets > 0:
                parts.append(f"[{skipped_sheets} additional sheet(s) not shown — sheet limit reached]")

            return "\n\n".join(parts) if parts else "(Excel file had no data)"
        except Exception as e:
            logger.debug(f"Excel parsing failed: {e}")
            return None

    @staticmethod
    def parse_pdf(data: bytes, max_pages: int = 20) -> str | None:
        """
        Parse PDF bytes and return extracted text with page markers, or None on parse failure.

        Extracts text from up to max_pages pages. Scanned/image-only PDFs
        will return an informational message (OCR is not performed).
        Password-protected PDFs return a clear message rather than a generic error.

        Args:
            data: Raw PDF file bytes
            max_pages: Maximum number of pages to extract

        Returns:
            Plain text with [Page N] markers (with truncation note if needed),
            or None on parse error
        """
        if not HAS_PYMUPDF:
            logger.debug("PDF parsing unavailable: pymupdf not installed")
            return None

        try:
            doc = fitz.open(stream=data, filetype="pdf")
            try:
                if doc.is_encrypted:
                    return "(PDF is password-protected and cannot be read without credentials)"

                total_pages = len(doc)
                pages_to_read = min(total_pages, max_pages)

                parts = []
                for page_num in range(pages_to_read):
                    page = doc[page_num]
                    text = page.get_text()
                    if text.strip():
                        parts.append(f"[Page {page_num + 1}]\n{text.strip()}")

                if not parts:
                    return "(PDF had no extractable text — may be a scanned/image-only document)"

                result = "\n\n".join(parts)
                if total_pages > max_pages:
                    result += f"\n\n[Showing {max_pages} of {total_pages} pages]"
                return result
            finally:
                doc.close()
        except Exception as e:
            logger.debug(f"PDF parsing failed: {e}")
            return None

    @staticmethod
    def parse_json(text: str, max_chars: int = 15000) -> str | None:
        """
        Parse a JSON string and return a readable representation, or None on parse failure.

        For arrays: shows item count and first 50 items.
        For objects: shows key count and all keys with their values.
        Truncates to max_chars to avoid excessive token usage.

        Args:
            text: Raw JSON text
            max_chars: Maximum output characters

        Returns:
            Formatted JSON string (truncated if needed), or None on error
        """
        try:
            data = json.loads(text)
        except json.JSONDecodeError as e:
            logger.debug(f"JSON parsing failed: {e}")
            return None

        try:
            if isinstance(data, list):
                total = len(data)
                preview = json.dumps(data[:50], indent=2, ensure_ascii=False)
                header = f"[JSON array — {total} items]\n"
                result = header + preview
                if total > 50:
                    result += f"\n... [{total - 50} more items]"
            elif isinstance(data, dict):
                header = f"[JSON object — {len(data)} keys]\n"
                result = header + json.dumps(data, indent=2, ensure_ascii=False)
            else:
                result = json.dumps(data, indent=2, ensure_ascii=False)

            if len(result) > max_chars:
                result = result[:max_chars] + "\n...[truncated]"
            return result
        except Exception as e:
            logger.debug(f"JSON formatting failed: {e}")
            return None


@dataclass
class ExtractedContent:
    """Extracted article content."""

    url: str
    title: str | None
    text: str
    success: bool
    method: str
    word_count: int
    author: str | None = None
    date: str | None = None
    source: str | None = None


class HTMLContentExtractor:
    """
    Multi-backend HTML content extractor with site-specific configurations.
    """

    # Site-specific extraction configs
    DEFAULT_SITE_CONFIGS = {
        "nytimes.com": {
            "cookies": {"nyt-gdpr": "accept"},
            "headers": {"Referer": "https://www.google.com/"},
            "wait_time": 2,
            "selectors": ["article#story", "div.StoryBodyCompanionColumn"],
            "remove_selectors": ["div.ad", "div.comments", "div.toolbar"],
        },
        "reuters.com": {
            "cookies": {"reuters-gdpr": "accept"},
            "headers": {"Referer": "https://news.google.com/"},
            "wait_time": 1,
            "selectors": ["article.article-body", 'div[data-testid="Body"]'],
            "remove_selectors": ["div.article-header", "div.article-share"],
        },
        "washingtonpost.com": {
            "cookies": {"wp_gdpr": "accept"},
            "headers": {"Referer": "https://www.google.com/"},
            "wait_time": 2,
            "selectors": ["div.article-body", "article.main-content"],
            "remove_selectors": ["div.interstitial", "div.newsletter-inline-unit"],
        },
        "economist.com": {
            "selectors": [
                "article.article__body",
                "div.article__lead",
                "div.article__content",
                "p.article__body-text",
            ],
            "remove_selectors": [
                "div.advert",
                "div.article__footnote",
                "div.share-links",
                "div.newsletter-signup",
            ],
        },
        "bloomberg.com": {
            "selectors": ["div.body-content", "div.body-copy", "article.article-body"],
            "remove_selectors": ["div.paywall", "div.newsletter-signup", "div.related-articles"],
        },
        "channelnewsasia.com": {
            "selectors": [
                "div.article-content",
                "div.article-body",
                "article",
                "div.article-text",
                "div.text-long",
            ],
            "remove_selectors": [
                "div.advertisement",
                "div.teaser",
                "div.partner-recommendations",
                "div.social-share",
                "div.also-worth-reading",
                "div.related-topics",
            ],
        },
        "techcrunch.com": {
            "selectors": [
                "div.article-content",
                "p.wp-block-paragraph",
                "ul.wp-block-list",
                "li.wp-block-list-item",
            ],
            "remove_selectors": [
                'div[class*="ad"]',
                'div[class*="related"]',
                'div[class*="promo"]',
                "aside",
                "nav",
                "header",
            ],
        },
    }

    # Patterns to remove from extracted content
    BLACKLIST_PATTERNS = [
        r"Advertisement\s*",
        r"ADVERTISEMENT\s*",
        r"Recommended\s*",
        r"###Embeddable###",
        r"Subscribe.*?(?=\n|$)",
        r"Sign up for.*?(?=\n|$)",
        r"Sign up here\.?",
        r"Newsletter.*?(?=\n|$)",
        r"Support quality journalism.*?(?=\n|$)",
        r"©\s*\d{4}.*?(?=\n|$)",
        r"All rights reserved.*?(?=\n|$)",
        r"PUBLISHED.*?, \d{2}:\d{2}.*?(?=\n|$)",
        r"PHOTO:.*?(?=\n|$)",
        r"Follow us on.*?(?=\n|$)",
        r"Share this article.*?(?=\n|$)",
        r"Thanks for sharing!.*?(?=\n|$)",
        r"Share full article.*?(?=\n|$)",
        r"Read more:.*?(?=\n|$)",
        r"More on this Topic.*?(?=\n|$)",
        r"See more on.*?(?=\n|$)",
        r"\[\s*\d+\s*chars\s*\]",
        r"^\s*https?://[^\s]+$",
    ]

    REMOVAL_KEYWORDS = [
        "embeddable",
        "subscribe",
        "subscription",
        "newsletter",
        "sign up",
        "share this",
        "follow us",
        "copyright",
        "all rights reserved",
        "read more",
        "more on this",
        "see more",
        "gallery",
        "slideshow",
        "click here",
        "download",
        "register",
        "privacy policy",
        "terms of service",
        "thanks for sharing",
        "photo:",
        "image:",
        "published",
        "updated",
    ]

    def __init__(self, site_configs: dict[str, dict] | None = None):
        self.site_configs = site_configs or self.DEFAULT_SITE_CONFIGS
        if HAS_BOILERPY:
            self.boilerpy_extractor = extractors.ArticleExtractor()
        else:
            self.boilerpy_extractor = None

    def _preprocess_html(self, html_content: str) -> str:
        """Pre-process HTML to improve extraction results."""
        try:
            soup = BeautifulSoup(html_content, "html.parser")

            # Remove script, style, and SVG tags
            for tag in soup.find_all(["script", "style", "svg", "noscript"]):
                tag.decompose()

            # Remove hidden elements
            for tag in soup.find_all(style=re.compile(r"display:\s*none|visibility:\s*hidden")):
                tag.decompose()

            # Remove footer/banner/sidebar/comment elements
            for div in soup.find_all("div", {"id": re.compile(r"footer|banner|sidebar|comment")}):
                div.decompose()

            # Remove header if it's navigation-like
            for div in soup.find_all("div", {"id": re.compile(r"header")}):
                header_text = div.get_text()
                if len(header_text) < 200 or re.search(
                    r"(menu|navigation|logo|sign in)", header_text.lower()
                ):
                    div.decompose()

            return str(soup)
        except Exception:
            return html_content

    def extract(self, url: str, html_content: str) -> tuple[str | None, str | None]:
        """
        Extract article content from HTML using multiple backends.

        Returns a tuple of (content, method) where method is one of:
        "site-specific", "trafilatura", "readability", "boilerpy",
        "density", "fallback", or None if extraction failed.
        """
        if not html_content or len(html_content.strip()) < 100:
            return None, None

        html_content = self._preprocess_html(html_content)
        metadata = self._get_article_metadata(html_content, url)
        soup = BeautifulSoup(html_content, "html.parser")

        cleaned_results = []

        # Strategy 1: Site-specific selectors
        site_specific = self._extract_with_selectors(html_content, url)
        if site_specific and len(site_specific.strip()) > 500:
            cleaned_results.append((site_specific, 1.2, "site-specific"))

        # Strategy 2: Trafilatura
        if HAS_TRAFILATURA:
            trafilatura_result = self._extract_trafilatura(html_content)
            if trafilatura_result:
                cleaned_results.append((trafilatura_result, 1.0, "trafilatura"))

        # Strategy 3: Readability
        if HAS_READABILITY:
            try:
                doc = Document(html_content)
                readability_soup = BeautifulSoup(doc.summary(), "html.parser")
                readability_text = self._fallback_extract_paragraphs(readability_soup)
                if readability_text:
                    cleaned_results.append((readability_text, 0.9, "readability"))
            except Exception as e:
                logger.debug(f"Readability failed: {e}")

        # Strategy 4: BoilerPy3
        if HAS_BOILERPY and self.boilerpy_extractor:
            try:
                boilerpy_result = self._extract_boilerpy(html_content)
                if boilerpy_result:
                    cleaned_results.append((boilerpy_result, 0.8, "boilerpy"))
            except Exception:
                pass

        # Score and select best result
        if cleaned_results:
            scored_results = []
            for content, base_weight, label in cleaned_results:
                score = self._calculate_content_quality(content)
                length_score = min(len(content) / 5000, 1.0)
                total_score = base_weight * score * length_score
                scored_results.append((content, total_score, label))

            best_result, _, label = max(scored_results, key=lambda scored: scored[1])
            logger.debug(f"Selected content from {label} with length {len(best_result)}")
            return self._format_with_metadata(self._clean_content(best_result), metadata), label

        # Fallback: auto-detect main content div
        guessed_div = self._find_content_by_density(soup)
        if guessed_div:
            text = guessed_div.get_text(separator=" ", strip=True)
            if text:
                return self._format_with_metadata(self._clean_content(text), metadata), "density"

        # Last resort: collect all <p> and <li> elements
        fallback = self._fallback_extract_paragraphs(soup)
        if fallback:
            return self._format_with_metadata(self._clean_content(fallback), metadata), "fallback"

        return None, None

    def _extract_with_selectors(self, html_content: str, url: str) -> str | None:
        """Extract content using site-specific CSS selectors."""
        domain = urlparse(url).netloc
        config = {}

        for known_domain in self.site_configs:
            if known_domain in domain:
                config = self.site_configs[known_domain]
                break

        selectors = config.get("selectors", [])
        remove_selectors = config.get("remove_selectors", [])

        if not selectors:
            selectors = [
                "article",
                ".article",
                ".article-body",
                ".content",
                ".entry-content",
                'div[itemprop="articleBody"]',
                ".story-body",
                ".post-content",
                "main",
            ]

        try:
            soup = BeautifulSoup(html_content, "html.parser")

            # Remove unwanted elements
            for selector in remove_selectors:
                for element in soup.select(selector):
                    element.decompose()

            for tag in soup.find_all(["aside", "nav", "footer"]):
                tag.decompose()

            for tag in soup.find_all(
                class_=re.compile(r"(ad-|banner|promo|sponsored|recommendation)")
            ):
                tag.decompose()

            content_parts = []

            for selector in selectors:
                for element in soup.select(selector):
                    self._clean_element(element)

                    paragraphs = []
                    children = list(element.children)
                    i = 0
                    while i < len(children):
                        child = children[i]
                        if not isinstance(child, Tag):
                            i += 1
                            continue

                        tag_name = child.name
                        text = child.get_text(separator=" ", strip=True)

                        if tag_name in ["h1", "h2", "h3", "h4", "h5", "h6"] and text:
                            merged_text = [text]
                            j = i + 1
                            while j < len(children):
                                next_child = children[j]
                                if isinstance(next_child, Tag) and next_child.name in [
                                    "p",
                                    "div",
                                    "ul",
                                    "ol",
                                ]:
                                    para = next_child.get_text(separator=" ", strip=True)
                                    if para and len(para) > 30 and not self._is_boilerplate(para):
                                        merged_text.append(para)
                                        i = j
                                        break
                                j += 1
                            paragraphs.append("\n".join(merged_text))
                        elif (
                            tag_name in ["p", "li"]
                            and text
                            and len(text) > 40
                            and not self._is_boilerplate(text)
                        ):
                            paragraphs.append(text)
                        elif tag_name == "table":
                            rows = child.find_all("tr")
                            table_lines = []
                            for row in rows:
                                cells = row.find_all(["th", "td"])
                                cell_texts = [c.get_text(strip=True) for c in cells]
                                table_lines.append("| " + " | ".join(cell_texts) + " |")
                                if row.find("th"):
                                    table_lines.append(
                                        "| " + " | ".join("---" for _ in cells) + " |"
                                    )
                            if table_lines:
                                paragraphs.append("\n".join(table_lines))
                        i += 1

                    if paragraphs:
                        content_parts.append("\n\n".join(paragraphs))

            if not content_parts:
                paragraphs = []
                for p in soup.find_all(["p"]):
                    text = p.get_text(separator=" ", strip=True)
                    if text and len(text) > 20 and not self._is_boilerplate(text):
                        paragraphs.append(text)
                if paragraphs:
                    return "\n\n".join(paragraphs)
                return None

            return "\n\n".join(content_parts)
        except Exception as e:
            logger.debug(f"Error in _extract_with_selectors: {e}")
            return None

    def _clean_element(self, element):
        """Clean an element's internal structure."""
        for tag in element.find_all(["script", "style", "button", "form", "input", "iframe"]):
            tag.decompose()

        for tag in element.find_all(
            class_=re.compile(r"share|social|comment|related|promo|ad|subscribe|newsletter")
        ):
            tag.decompose()

        for p in element.find_all("p"):
            if (
                not p.get_text(separator=" ", strip=True)
                or len(p.get_text(separator=" ", strip=True)) < 5
            ):
                p.decompose()

    def _is_boilerplate(self, text: str) -> bool:
        """Check if text is likely boilerplate content."""
        text_lower = text.lower()

        if any(keyword in text_lower for keyword in self.REMOVAL_KEYWORDS):
            return True

        if len(text) < 20 and (
            text.isupper()
            or text.count(".") == 0
            or re.search(r"^\d+:\d+$", text)
            or re.search(r"^[^\w]*https?://", text)
        ):
            return True

        return False

    def _extract_trafilatura(self, html_content: str) -> str | None:
        """Extract content using trafilatura."""
        if not HAS_TRAFILATURA:
            return None
        try:
            config = use_config()
            if not config.has_section("EXTRACTION"):
                config.add_section("EXTRACTION")
            config.set("DEFAULT", "EXTRACTION_TIMEOUT", "0")
            config.set("EXTRACTION", "favor_precision", "true")

            result = trafilatura_extract(html_content, config=config)
            return result if result and len(result) > 300 else None
        except Exception as e:
            logger.debug(f"Trafilatura failed: {e}")
            return None

    def _extract_boilerpy(self, html_content: str) -> str | None:
        """Extract content using boilerpy3."""
        if not self.boilerpy_extractor:
            return None
        try:
            content = self.boilerpy_extractor.get_content(html_content)
            if content:
                lines = content.split("\n")
                filtered_lines = [
                    line for line in lines if len(line) > 20 and not self._is_boilerplate(line)
                ]
                return "\n\n".join(filtered_lines) if filtered_lines else None
            return None
        except Exception:
            return None

    def _find_content_by_density(self, soup: BeautifulSoup) -> Tag | None:
        """Find the main content div based on text density."""
        candidates = []
        for div in soup.find_all("div"):
            class_attr = " ".join(div.get("class", []))
            if any(
                kw in class_attr for kw in ["footer", "nav", "sidebar", "header", "promo", "share"]
            ):
                continue
            text = div.get_text(separator=" ", strip=True)
            p_count = len(div.find_all("p"))
            if len(text) > 500 and p_count >= 3:
                candidates.append((div, len(text)))

        if candidates:
            return sorted(candidates, key=lambda x: x[1], reverse=True)[0][0]
        return None

    def _fallback_extract_paragraphs(self, soup: BeautifulSoup) -> str | None:
        """Fallback: extract all meaningful paragraphs."""
        paragraphs = []
        for tag in soup.find_all(["p", "li", "h2", "h3"]):
            text = tag.get_text(separator=" ", strip=True)
            if text and len(text) > 40 and not self._is_boilerplate(text):
                paragraphs.append(text)
        return "\n\n".join(paragraphs) if paragraphs else None

    def _calculate_content_quality(self, content: str) -> float:
        """Calculate a quality score for the content."""
        if not content:
            return 0.0

        avg_line_length = sum(len(line) for line in content.splitlines()) / max(
            1, len(content.splitlines())
        )
        sentence_count = content.count(".") + content.count("!") + content.count("?")
        word_count = len(content.split())

        quality_score = min(
            1.0, (sentence_count / max(1, word_count / 20)) * (min(avg_line_length, 100) / 100)
        )

        alphanumeric_ratio = sum(c.isalnum() or c.isspace() for c in content) / max(1, len(content))
        quality_score *= max(0.5, alphanumeric_ratio)

        uppercase_ratio = sum(c.isupper() for c in content) / max(
            1, len(content) - content.count(" ")
        )
        if uppercase_ratio > 0.3:
            quality_score *= 0.8

        return quality_score

    def _clean_content(self, content: str) -> str:
        """Clean extracted content."""
        if not content:
            return ""

        content = html_lib.unescape(content)

        # Remove invisible/control characters
        content = re.sub(r"[\x00-\x1F\x7F]", " ", content)

        # Fix punctuation spacing
        content = re.sub(r"\s+([.,;!?])", r"\1", content)
        content = re.sub(r"([.,;!?])(?=\w)", r"\1 ", content)

        # Fix joined words
        content = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", content)
        content = re.sub(r"(?<=[a-zA-Z])(?=\d)", " ", content)
        content = re.sub(r"(?<=\d)(?=[a-zA-Z])", " ", content)

        # Normalize spacing
        content = re.sub(r" {2,}", " ", content)
        content = re.sub(r"\n{3,}", "\n\n", content)

        # Apply blacklist patterns
        for pattern in self.BLACKLIST_PATTERNS:
            content = re.sub(pattern, " ", content, flags=re.IGNORECASE | re.MULTILINE)

        # Final cleanup
        content = re.sub(r"\n{3,}", "\n\n", content)
        content = content.strip()

        return content

    def _get_article_metadata(self, html_content: str, url: str) -> dict:
        """Extract metadata from the article."""
        try:
            metadata = {}
            soup = BeautifulSoup(html_content, "html.parser")

            metadata["title"] = self._extract_title(soup)
            metadata["author"] = self._extract_author(soup)
            metadata["date"] = self._extract_date(soup)
            metadata["source"] = self._extract_source(soup, url)

            return {k: v for k, v in metadata.items() if v}
        except Exception:
            return {}

    def _extract_title(self, soup: BeautifulSoup) -> str | None:
        """Extract the article title."""
        for meta in soup.find_all("meta"):
            if meta.get("property") in ["og:title", "twitter:title"] and meta.get("content"):
                return meta.get("content").strip()

        h1 = soup.find("h1")
        if h1:
            return h1.get_text(separator=" ", strip=True)

        title_tag = soup.find("title")
        if title_tag:
            return title_tag.get_text(separator=" ", strip=True)

        return None

    def _extract_author(self, soup: BeautifulSoup) -> str | None:
        """Extract the author name."""
        for meta in soup.find_all("meta"):
            if meta.get("name") in ["author", "article:author"] and meta.get("content"):
                return meta.get("content").strip()

        for selector in [".author", ".byline", ".writer", '[rel="author"]']:
            author_elem = soup.select_one(selector)
            if author_elem:
                return author_elem.get_text(separator=" ", strip=True)

        return None

    def _extract_date(self, soup: BeautifulSoup) -> str | None:
        """Extract publication date."""
        for meta in soup.find_all("meta"):
            if meta.get("property") in ["article:published_time", "og:published_time"] and meta.get(
                "content"
            ):
                return meta.get("content").strip()

        time_elem = soup.find("time")
        if time_elem and time_elem.get("datetime"):
            return time_elem.get("datetime")

        for selector in [".date", ".published", ".timestamp", ".pubdate"]:
            date_elem = soup.select_one(selector)
            if date_elem:
                return date_elem.get_text(separator=" ", strip=True)

        return None

    def _extract_source(self, soup: BeautifulSoup, url: str) -> str:
        """Extract publication source name."""
        for meta in soup.find_all("meta"):
            if meta.get("property") in ["og:site_name"] and meta.get("content"):
                return meta.get("content").strip()

        domain = urlparse(url).netloc
        source = domain.replace("www.", "")
        source = re.sub(r"\.(com|org|net|gov|edu|io)$", "", source)
        return source.capitalize()

    def _format_with_metadata(self, content: str, metadata: dict) -> str:
        """Format content with metadata header."""
        if not content:
            return ""

        has_title = False
        if "title" in metadata and metadata["title"]:
            title_start = metadata["title"][:30].lower()
            content_start = content[:100].lower()
            has_title = title_start in content_start

        header_parts = []

        if "title" in metadata and metadata["title"] and not has_title:
            header_parts.append(f"# {metadata['title']}")

        meta_parts = []

        if "source" in metadata and metadata["source"]:
            meta_parts.append(f"Content source: {metadata['source']}")

        if "date" in metadata and metadata["date"]:
            try:
                from dateutil.parser import parse as parse_date

                parsed_date = parse_date(metadata["date"])
                formatted_date = parsed_date.strftime("%B %d, %Y")
                meta_parts.append(f"Date: {formatted_date}")
            except Exception:
                meta_parts.append(f"Date: {metadata['date']}")

        if "author" in metadata and metadata["author"]:
            meta_parts.append(f"Author: {metadata['author']}")

        if meta_parts:
            header_parts.append("\n".join(meta_parts))

        if header_parts:
            return "\n\n".join(header_parts + [content])

        return content


class ConcurrentContentExtractor:
    """
    Concurrent URL fetcher and content extractor.

    Uses httpx for async HTTP requests and HTMLContentExtractor for parsing.
    Optionally supports Bright Data API for more reliable scraping.
    """

    # Sites that commonly block or have anti-bot measures
    BLOCKED_DOMAINS = [
        "twitter.com",
        "x.com",
        "facebook.com",
        "fb.com",
        "instagram.com",
        "linkedin.com",
        "tiktok.com",
        "trends.google.com",  # JS-rendered app; bot has dedicated Google Trends integration
    ]

    # Maximum content length to return (chars)
    MAX_CONTENT_LENGTH = 15000

    # Request timeout
    TIMEOUT = 15.0

    def __init__(self, bright_data_api_key: str | None = None):
        """
        Initialize the concurrent content extractor.

        Args:
            bright_data_api_key: Optional Bright Data API key for more reliable scraping.
                                If not provided, uses direct HTTP requests.
        """
        self.bright_data_api_key = bright_data_api_key
        self.html_extractor = HTMLContentExtractor()
        self._client: httpx.AsyncClient | None = None

    async def __aenter__(self):
        self._client = httpx.AsyncClient(
            timeout=self.TIMEOUT,
            follow_redirects=True,
            headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.5",
            },
        )
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        if self._client:
            await self._client.aclose()
            self._client = None

    def _is_blocked_domain(self, url: str) -> bool:
        """Check if URL is from a domain that typically blocks scraping."""
        for domain in self.BLOCKED_DOMAINS:
            if domain in url.lower():
                return True
        return False

    @staticmethod
    def _is_wikipedia_url(url: str) -> bool:
        """Check if URL is a Wikipedia page."""
        return "wikipedia.org/wiki/" in url.lower()

    @staticmethod
    def _is_google_sheets_url(url: str) -> bool:
        """Check if URL points to a Google Sheets document."""
        parsed = urlparse(url)
        return parsed.netloc.lower() == "docs.google.com" and parsed.path.lower().startswith(
            "/spreadsheets/"
        )

    @staticmethod
    def _rewrite_google_sheets_url(url: str) -> str:
        """
        Rewrite a Google Sheets URL to its direct CSV export equivalent.

        Extracts the spreadsheet ID from the URL path and the sheet tab ID (gid)
        from the URL fragment. Falls back to gid=0 (first sheet) if absent.

        Examples:
            https://docs.google.com/spreadsheets/d/ABC/edit#gid=123
            -> https://docs.google.com/spreadsheets/d/ABC/export?format=csv&gid=123

            https://docs.google.com/spreadsheets/d/ABC/pub
            -> https://docs.google.com/spreadsheets/d/ABC/export?format=csv&gid=0
        """
        parsed = urlparse(url)
        # Extract spreadsheet ID: path looks like /spreadsheets/d/{ID}/...
        path_parts = parsed.path.rstrip("/").split("/")
        try:
            d_index = path_parts.index("d")
            spreadsheet_id = path_parts[d_index + 1]
        except (ValueError, IndexError):
            return url  # Can't parse — return unchanged and let it fail naturally

        # Extract gid from fragment (e.g. "#gid=123") or query params (e.g. "?gid=123")
        gid = "0"
        for source in (parsed.fragment, parsed.query):
            if source:
                gid_match = re.search(r"gid=(\d+)", source)
                if gid_match:
                    gid = gid_match.group(1)
                    break

        return (
            f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv&gid={gid}"
        )

    _WIKIPEDIA_USER_AGENT = "MetaculusBot/1.0 (https://github.com/metaculus; forecasting bot)"

    @staticmethod
    def _make_result(
        url: str,
        domain: str,
        *,
        content: str | None = None,
        success: bool = False,
        error: str | None = None,
        status_code: int | None = None,
        method: str | None = None,
        content_chars: int = 0,
        truncated: bool = False,
    ) -> dict[str, Any]:
        """Build a standardized extraction result dict."""
        return {
            "url": url,
            "domain": domain,
            "content": content,
            "success": success,
            "error": error,
            "status_code": status_code,
            "method": method,
            "content_chars": content_chars,
            "truncated": truncated,
        }

    async def _fetch_wikipedia(self, url: str) -> dict[str, Any]:
        """Fetch Wikipedia content via the MediaWiki API.

        Uses action=query&prop=extracts for article pages (returns clean
        plain text) and action=query&meta=siteinfo for Special:Statistics.
        """
        parsed = urlparse(url)
        domain = parsed.netloc
        path = unquote(parsed.path)

        if "/wiki/" not in path:
            return await self._fetch_url(url)

        page_title = path.split("/wiki/", 1)[1]

        # Strip fragment (e.g. #Launch_date)
        if "#" in page_title:
            page_title = page_title.split("#", 1)[0]

        if not page_title:
            return self._make_result(
                url, domain, error="Empty Wikipedia page title", method="wikipedia_api"
            )

        try:
            api_base = f"https://{domain}/w/api.php"
            headers = {"User-Agent": self._WIKIPEDIA_USER_AGENT}

            # Special:Statistics -> use siteinfo API
            if page_title == "Special:Statistics":
                response = await self._client.get(
                    api_base,
                    params={
                        "action": "query",
                        "meta": "siteinfo",
                        "siprop": "statistics",
                        "format": "json",
                    },
                    headers=headers,
                )
                response.raise_for_status()
                data = response.json()

                stats = data.get("query", {}).get("statistics", {})
                if not stats:
                    return self._make_result(
                        url,
                        domain,
                        error="No statistics data returned",
                        status_code=response.status_code,
                        method="wikipedia_api",
                    )

                def _fmt(val: int | str) -> str:
                    return f"{val:,}" if isinstance(val, int) else str(val)

                content = "# Wikipedia Statistics\n\n"
                content += f"Content pages (articles): {_fmt(stats.get('articles', 'N/A'))}\n"
                content += f"Total pages: {_fmt(stats.get('pages', 'N/A'))}\n"
                content += f"Total edits: {_fmt(stats.get('edits', 'N/A'))}\n"
                content += f"Uploaded files: {_fmt(stats.get('images', 'N/A'))}\n"
                content += f"Registered users: {_fmt(stats.get('users', 'N/A'))}\n"
                content += f"Active users: {_fmt(stats.get('activeusers', 'N/A'))}\n"
                content += f"Administrators: {_fmt(stats.get('admins', 'N/A'))}\n"

                return self._make_result(
                    url,
                    domain,
                    content=content,
                    success=True,
                    status_code=response.status_code,
                    method="wikipedia_api",
                    content_chars=len(content),
                )

            # Regular article -> use extracts API
            response = await self._client.get(
                api_base,
                params={
                    "action": "query",
                    "titles": page_title,
                    "prop": "extracts",
                    "explaintext": 1,
                    "format": "json",
                },
                headers=headers,
            )
            response.raise_for_status()
            data = response.json()

            pages = data.get("query", {}).get("pages", {})
            if not pages:
                return self._make_result(
                    url,
                    domain,
                    error="No pages in API response",
                    status_code=response.status_code,
                    method="wikipedia_api",
                )

            page_data = next(iter(pages.values()))

            if "missing" in page_data:
                return self._make_result(
                    url,
                    domain,
                    error=f"Wikipedia page not found: {page_data.get('title', page_title)}",
                    status_code=response.status_code,
                    method="wikipedia_api",
                )

            extract = page_data.get("extract", "")
            if not extract or len(extract.strip()) < 50:
                return self._make_result(
                    url,
                    domain,
                    error="Wikipedia extract too short",
                    status_code=response.status_code,
                    method="wikipedia_api",
                )

            title = page_data.get("title", page_title.replace("_", " "))
            content = f"# {title}\n\n{extract}"

            truncated = len(content) > self.MAX_CONTENT_LENGTH
            content_chars = len(content)
            if truncated:
                content = content[: self.MAX_CONTENT_LENGTH] + "...[truncated]"

            return self._make_result(
                url,
                domain,
                content=content,
                success=True,
                status_code=response.status_code,
                method="wikipedia_api",
                content_chars=content_chars,
                truncated=truncated,
            )

        except Exception as e:
            logger.debug(f"Wikipedia API failed for {url}: {e}")
            return self._make_result(
                url, domain, error=f"Wikipedia API error: {e}", method="wikipedia_api"
            )

    async def _fetch_url(self, url: str) -> dict[str, Any]:
        """Fetch a single URL and extract content."""
        domain = urlparse(url).netloc
        if self._is_blocked_domain(url):
            return self._make_result(url, domain, error="Blocked domain")

        if self._is_wikipedia_url(url):
            return await self._fetch_wikipedia(url)

        # Google Sheets: rewrite to CSV export URL before fetching
        fetch_url = url
        if self._is_google_sheets_url(url):
            fetch_url = self._rewrite_google_sheets_url(url)
            logger.info(f"[spreadsheet] Rewriting Google Sheets URL to CSV export: {fetch_url}")

        # Detect non-HTML by extension before fetching (Bright Data is for web pages only)
        url_ext = url.lower().split("?")[0]
        is_non_html_by_extension = url_ext.endswith((".csv", ".xlsx", ".xls", ".pdf", ".json"))

        try:
            # Skip Bright Data for non-HTML content (spreadsheets, PDFs, JSON, Google Sheets)
            if self.bright_data_api_key and not self._is_google_sheets_url(url) and not is_non_html_by_extension:
                return await self._fetch_with_bright_data(url)

            # Direct HTTP request
            response = await self._client.get(fetch_url)
            status_code = response.status_code
            response.raise_for_status()

            # Detect spreadsheet content by Content-Type header or URL extension
            content_type = response.headers.get("content-type", "").lower()
            url_lower = fetch_url.lower().split("?")[0]  # strip query params for extension check

            is_csv = "text/csv" in content_type or url_lower.endswith(".csv")
            is_excel = (
                "spreadsheetml" in content_type
                or "ms-excel" in content_type
                or url_lower.endswith(".xlsx")
                or url_lower.endswith(".xls")
            )
            is_pdf = "application/pdf" in content_type or url_lower.endswith(".pdf")
            is_json = (
                "application/json" in content_type
                or "text/json" in content_type
                or url_lower.endswith(".json")
            )
            # Google Sheets export always returns CSV regardless of content-type header
            is_google_sheets = self._is_google_sheets_url(url)

            if is_csv or is_google_sheets:
                parsed_content = SpreadsheetParser.parse_csv(response.text)
                if parsed_content is None:
                    return self._make_result(
                        url, domain, success=False, error="CSV parsing failed", status_code=status_code, method="spreadsheet_csv"
                    )
                return self._make_result(
                    url, domain, content=parsed_content, success=True, status_code=status_code,
                    method="spreadsheet_csv", content_chars=len(parsed_content),
                )

            if is_excel:
                parsed_content = SpreadsheetParser.parse_excel(response.content)
                if parsed_content is None:
                    return self._make_result(
                        url, domain, success=False, error="Excel parsing failed", status_code=status_code, method="spreadsheet_excel"
                    )
                return self._make_result(
                    url, domain, content=parsed_content, success=True, status_code=status_code,
                    method="spreadsheet_excel", content_chars=len(parsed_content),
                )

            if is_pdf:
                parsed_content = SpreadsheetParser.parse_pdf(response.content)
                if parsed_content is None:
                    return self._make_result(
                        url, domain, success=False, error="PDF parsing failed", status_code=status_code, method="pdf"
                    )
                return self._make_result(
                    url, domain, content=parsed_content, success=True, status_code=status_code,
                    method="pdf", content_chars=len(parsed_content),
                )

            if is_json:
                parsed_content = SpreadsheetParser.parse_json(response.text)
                if parsed_content is None:
                    return self._make_result(
                        url, domain, success=False, error="JSON parsing failed", status_code=status_code, method="json"
                    )
                return self._make_result(
                    url, domain, content=parsed_content, success=True, status_code=status_code,
                    method="json", content_chars=len(parsed_content),
                )

            raw_html = response.text

            if not raw_html or len(raw_html.strip()) < 100:
                return self._make_result(
                    url,
                    domain,
                    error="Empty or very short HTML received",
                    status_code=status_code,
                )

            # Extract content
            processed_content, method = self.html_extractor.extract(url, raw_html)

            if not processed_content:
                return self._make_result(
                    url,
                    domain,
                    error="Content extraction failed",
                    status_code=status_code,
                )

            # Truncate if too long
            truncated = len(processed_content) > self.MAX_CONTENT_LENGTH
            content_chars = len(processed_content)
            if truncated:
                processed_content = processed_content[: self.MAX_CONTENT_LENGTH] + "...[truncated]"

            return self._make_result(
                url,
                domain,
                content=processed_content,
                success=True,
                status_code=status_code,
                method=method,
                content_chars=content_chars,
                truncated=truncated,
            )

        except (TimeoutError, httpx.TimeoutException):
            return self._make_result(url, domain, error="Request timed out")
        except httpx.HTTPStatusError as e:
            logger.debug(f"HTTP error for {url}: {e}")
            return self._make_result(
                url,
                domain,
                error=f"HTTP {e.response.status_code}",
                status_code=e.response.status_code,
            )
        except Exception as e:
            logger.debug(f"Error processing {url}: {e}")
            return self._make_result(url, domain, error=str(e))

    async def _fetch_with_bright_data(self, url: str) -> dict[str, Any]:
        """Fetch URL using Bright Data API."""
        try:
            headers = {
                "Authorization": f"Bearer {self.bright_data_api_key}",
                "Content-Type": "application/json",
            }
            payload = {
                "url": url,
                "zone": "web_scraper",
                "format": "raw",
            }

            response = await self._client.post(
                "https://api.brightdata.com/request", headers=headers, json=payload
            )

            domain = urlparse(url).netloc

            if response.status_code != 200:
                return self._make_result(
                    url,
                    domain,
                    error=f"Bright Data API error: {response.status_code}",
                    status_code=response.status_code,
                )

            raw_html = response.text

            if not raw_html or len(raw_html.strip()) < 100:
                return self._make_result(
                    url, domain, error="Empty HTML from Bright Data", status_code=200
                )

            processed_content, method = self.html_extractor.extract(url, raw_html)

            if not processed_content:
                return self._make_result(
                    url, domain, error="Content extraction failed", status_code=200
                )

            truncated = len(processed_content) > self.MAX_CONTENT_LENGTH
            content_chars = len(processed_content)
            if truncated:
                processed_content = processed_content[: self.MAX_CONTENT_LENGTH] + "...[truncated]"

            return self._make_result(
                url,
                domain,
                content=processed_content,
                success=True,
                status_code=200,
                method=method,
                content_chars=content_chars,
                truncated=truncated,
            )

        except Exception as e:
            logger.debug(f"Bright Data error for {url}: {e}")
            return self._make_result(url, urlparse(url).netloc, error=str(e))

    async def extract_content(self, urls: list[str]) -> dict[str, dict[str, Any]]:
        """
        Extract content from multiple URLs concurrently.

        Args:
            urls: List of URLs to extract

        Returns:
            Dict mapping URL to extraction result
        """
        results = {}

        try:
            tasks = [asyncio.create_task(self._fetch_url(url)) for url in urls]

            # Wait for tasks with timeout
            done, pending = await asyncio.wait(tasks, timeout=75)

            # Handle completed tasks
            for task in done:
                try:
                    result = task.result()
                    results[result["url"]] = result
                except Exception as e:
                    logger.debug(f"Error getting task result: {e}")

            # Handle pending tasks (timed out)
            for task in pending:
                task.cancel()
                try:
                    task_index = None
                    for i, t in enumerate(tasks):
                        if t == task:
                            task_index = i
                            break

                    url = urls[task_index] if task_index is not None else "unknown"
                    logger.debug(f"Task for {url} timed out")
                    results[url] = {
                        "url": url,
                        "domain": urlparse(url).netloc if url != "unknown" else "unknown",
                        "content": None,
                        "error": "Operation timed out",
                        "success": False,
                    }
                except Exception as e:
                    logger.debug(f"Error handling cancelled task: {e}")

        except Exception as e:
            logger.error(f"Error in extract_content: {e}")

        return results

    async def process_urls(self, urls: list[str]) -> dict[str, tuple[str, bool]]:
        """
        Process URLs and return simplified results.

        Args:
            urls: List of URLs to process

        Returns:
            Dict mapping URL to (content, success) tuple
        """
        results_dict = await self.extract_content(urls)
        processed_results = {}

        for url, result in results_dict.items():
            content = result.get("content", "")
            success = result.get("success", False)
            processed_results[url] = (content or "", success)

        return processed_results


async def extract_article(url: str) -> ExtractedContent:
    """
    Convenience function to extract a single article.

    Usage:
        content = await extract_article("https://example.com/article")
        print(content.text)
    """
    async with ConcurrentContentExtractor() as extractor:
        results = await extractor.extract_content([url])

        if url in results and results[url]["success"]:
            result = results[url]
            # Try to parse metadata from content
            content_text = result["content"] or ""
            title = None

            # Try to extract title from first line if it starts with #
            if content_text.startswith("# "):
                first_line_end = content_text.find("\n")
                if first_line_end > 0:
                    title = content_text[2:first_line_end].strip()

            return ExtractedContent(
                url=url,
                title=title,
                text=content_text,
                success=True,
                method="fast_extractor",
                word_count=len(content_text.split()),
            )

        return ExtractedContent(
            url=url, title=None, text="", success=False, method="failed", word_count=0
        )


async def extract_articles_batch(
    urls: list[str], max_concurrent: int = 10
) -> list[ExtractedContent]:
    """
    Extract multiple articles concurrently.

    Usage:
        urls = ["https://example.com/1", "https://example.com/2"]
        results = await extract_articles_batch(urls)
        for result in results:
            print(f"{result.url}: {result.word_count} words")
    """
    async with ConcurrentContentExtractor() as extractor:
        results_dict = await extractor.extract_content(urls)

        results = []
        for url in urls:
            if url in results_dict and results_dict[url]["success"]:
                result = results_dict[url]
                content_text = result["content"] or ""
                title = None

                if content_text.startswith("# "):
                    first_line_end = content_text.find("\n")
                    if first_line_end > 0:
                        title = content_text[2:first_line_end].strip()

                results.append(
                    ExtractedContent(
                        url=url,
                        title=title,
                        text=content_text,
                        success=True,
                        method="fast_extractor",
                        word_count=len(content_text.split()),
                    )
                )
            else:
                results.append(
                    ExtractedContent(
                        url=url, title=None, text="", success=False, method="failed", word_count=0
                    )
                )

        return results
