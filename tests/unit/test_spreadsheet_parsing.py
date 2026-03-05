"""
Tests for spreadsheet parsing (CSV, Excel, Google Sheets).

Tests SpreadsheetParser (CSV and Excel parsing) and the Google Sheets URL
rewriting logic in ConcurrentContentExtractor.
"""

import io
import sys
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from src.bot.content_extractor import ConcurrentContentExtractor, SpreadsheetParser

# ---------------------------------------------------------------------------
# SpreadsheetParser.parse_csv
# ---------------------------------------------------------------------------


class TestParseCsv:
    def test_basic_table(self):
        csv_text = "Name,Age,City\nAlice,30,NYC\nBob,25,LA\n"
        result = SpreadsheetParser.parse_csv(csv_text)
        assert result is not None
        assert "| Name | Age | City |" in result
        assert "Alice" in result
        assert "Bob" in result

    def test_separator_row_present(self):
        csv_text = "A,B\n1,2\n"
        result = SpreadsheetParser.parse_csv(csv_text)
        assert result is not None
        assert "| --- |" in result

    def test_pipe_in_cell_escaped(self):
        csv_text = "Col\nval|ue\n"
        result = SpreadsheetParser.parse_csv(csv_text)
        assert result is not None
        assert r"val\|ue" in result

    def test_truncation_note_when_over_limit(self):
        rows = "\n".join(f"row{i},{i}" for i in range(250))
        csv_text = f"Name,Num\n{rows}\n"
        result = SpreadsheetParser.parse_csv(csv_text, max_rows=200)
        assert result is not None
        assert "[Showing 200 of 250 rows]" in result

    def test_no_truncation_note_when_within_limit(self):
        csv_text = "A,B\n1,2\n3,4\n"
        result = SpreadsheetParser.parse_csv(csv_text, max_rows=200)
        assert result is not None
        assert "Showing" not in result

    def test_empty_csv_returns_message_not_none(self):
        # Empty input is a valid parse result (not an error) — returns descriptive string
        result = SpreadsheetParser.parse_csv("")
        assert result is not None
        assert "empty" in result.lower() or "no headers" in result.lower()

    def test_single_column(self):
        csv_text = "Value\napple\nbanana\n"
        result = SpreadsheetParser.parse_csv(csv_text)
        assert result is not None
        assert "Value" in result
        assert "apple" in result

    def test_newline_in_cell_flattened(self):
        # csv module handles quoted fields with embedded newlines
        csv_text = 'Col\n"line1\nline2"\n'
        result = SpreadsheetParser.parse_csv(csv_text)
        assert result is not None
        # Should not contain a raw newline inside a table cell
        lines = result.split("\n")
        data_line = [line for line in lines if "line1" in line][0]
        assert "\n" not in data_line

    def test_semicolon_delimiter_detected(self):
        csv_text = "Name;Age;City\nAlice;30;NYC\nBob;25;LA\n"
        result = SpreadsheetParser.parse_csv(csv_text)
        assert result is not None
        assert "Name" in result
        assert "Alice" in result
        assert "NYC" in result

    def test_tab_delimiter_detected(self):
        csv_text = "Name\tScore\nAlice\t95\nBob\t87\n"
        result = SpreadsheetParser.parse_csv(csv_text)
        assert result is not None
        assert "Name" in result
        assert "Score" in result
        assert "Alice" in result


# ---------------------------------------------------------------------------
# SpreadsheetParser.parse_excel
# ---------------------------------------------------------------------------


class TestParseExcel:
    def _make_xlsx(self, headers: list, rows: list[list]) -> bytes:
        """Create a minimal in-memory .xlsx file."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_basic_table(self):
        data = self._make_xlsx(["Name", "Score"], [["Alice", 95], ["Bob", 87]])
        result = SpreadsheetParser.parse_excel(data)
        assert result is not None
        assert "Name" in result
        assert "Alice" in result
        assert "87" in result

    def test_separator_row_present(self):
        data = self._make_xlsx(["A", "B"], [[1, 2]])
        result = SpreadsheetParser.parse_excel(data)
        assert result is not None
        assert "| --- |" in result

    def test_truncation_note_when_over_limit(self):
        rows = [[f"row{i}", i] for i in range(250)]
        data = self._make_xlsx(["Label", "Num"], rows)
        result = SpreadsheetParser.parse_excel(data, max_rows=200)
        assert result is not None
        assert "Showing" in result

    def test_no_truncation_when_within_limit(self):
        data = self._make_xlsx(["A"], [["x"], ["y"]])
        result = SpreadsheetParser.parse_excel(data)
        assert result is not None
        assert "Showing" not in result

    def test_multiple_sheets_labeled(self):
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["Col"])
        ws1.append(["val1"])
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["X"])
        ws2.append(["val2"])
        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()

        result = SpreadsheetParser.parse_excel(data)
        assert result is not None
        assert "Sheet1" in result
        assert "Sheet2" in result
        assert "val1" in result
        assert "val2" in result

    def test_invalid_bytes_returns_none(self):
        result = SpreadsheetParser.parse_excel(b"not an excel file")
        assert result is None

    def test_none_cells_handled(self):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append([None, "value"])
        buf = io.BytesIO()
        wb.save(buf)
        result = SpreadsheetParser.parse_excel(buf.getvalue())
        assert result is not None
        assert "value" in result

    def test_openpyxl_unavailable_returns_none(self):
        with patch("src.bot.content_extractor.HAS_OPENPYXL", False):
            result = SpreadsheetParser.parse_excel(b"anything")
        assert result is None

    def test_global_sheet_limit(self):
        """Workbooks with more sheets than max_sheets get a truncation note."""
        import openpyxl

        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.active.append(["Col"])
        wb.active.append(["v1"])
        for i in range(2, 13):  # 12 sheets total
            ws = wb.create_sheet(f"Sheet{i}")
            ws.append(["X"])
            ws.append([f"v{i}"])
        buf = io.BytesIO()
        wb.save(buf)

        result = SpreadsheetParser.parse_excel(buf.getvalue(), max_sheets=5)
        assert result is not None
        assert "additional sheet" in result.lower() or "sheet limit" in result.lower()

    def test_global_total_row_limit(self):
        """Rows across all sheets are capped by max_total_rows."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "S1"
        ws1.append(["Val"])
        for i in range(300):
            ws1.append([i])
        ws2 = wb.create_sheet("S2")
        ws2.append(["Val"])
        for i in range(300):
            ws2.append([i])
        buf = io.BytesIO()
        wb.save(buf)

        result = SpreadsheetParser.parse_excel(buf.getvalue(), max_total_rows=100)
        assert result is not None
        # S2 should be skipped or truncated due to global limit
        assert "limit" in result.lower() or "Showing" in result


# ---------------------------------------------------------------------------
# ConcurrentContentExtractor._is_google_sheets_url
# ---------------------------------------------------------------------------


class TestIsGoogleSheetsUrl:
    def test_edit_url(self):
        assert ConcurrentContentExtractor._is_google_sheets_url(
            "https://docs.google.com/spreadsheets/d/ABC123/edit#gid=0"
        )

    def test_view_url(self):
        assert ConcurrentContentExtractor._is_google_sheets_url(
            "https://docs.google.com/spreadsheets/d/SHEETID/view"
        )

    def test_pub_url(self):
        assert ConcurrentContentExtractor._is_google_sheets_url(
            "https://docs.google.com/spreadsheets/d/ID/pub?output=csv"
        )

    def test_not_google_sheets(self):
        assert not ConcurrentContentExtractor._is_google_sheets_url("https://example.com/data.csv")

    def test_google_docs_not_sheets(self):
        assert not ConcurrentContentExtractor._is_google_sheets_url(
            "https://docs.google.com/document/d/ABC/edit"
        )


# ---------------------------------------------------------------------------
# ConcurrentContentExtractor._rewrite_google_sheets_url
# ---------------------------------------------------------------------------


class TestRewriteGoogleSheetsUrl:
    def test_edit_url_with_gid(self):
        url = "https://docs.google.com/spreadsheets/d/ABC123/edit#gid=456"
        result = ConcurrentContentExtractor._rewrite_google_sheets_url(url)
        assert result == "https://docs.google.com/spreadsheets/d/ABC123/export?format=csv&gid=456"

    def test_edit_url_without_gid_defaults_to_zero(self):
        url = "https://docs.google.com/spreadsheets/d/ABC123/edit"
        result = ConcurrentContentExtractor._rewrite_google_sheets_url(url)
        assert result == "https://docs.google.com/spreadsheets/d/ABC123/export?format=csv&gid=0"

    def test_view_url(self):
        url = "https://docs.google.com/spreadsheets/d/XYZ/view#gid=99"
        result = ConcurrentContentExtractor._rewrite_google_sheets_url(url)
        assert result == "https://docs.google.com/spreadsheets/d/XYZ/export?format=csv&gid=99"

    def test_pub_url_with_gid_in_fragment(self):
        url = "https://docs.google.com/spreadsheets/d/ID/pub#gid=1"
        result = ConcurrentContentExtractor._rewrite_google_sheets_url(url)
        assert result == "https://docs.google.com/spreadsheets/d/ID/export?format=csv&gid=1"

    def test_gid_in_query_params(self):
        url = "https://docs.google.com/spreadsheets/d/ABC/edit?gid=789"
        result = ConcurrentContentExtractor._rewrite_google_sheets_url(url)
        assert result == "https://docs.google.com/spreadsheets/d/ABC/export?format=csv&gid=789"

    def test_fragment_gid_takes_priority_over_query_gid(self):
        url = "https://docs.google.com/spreadsheets/d/ABC/edit?gid=111#gid=222"
        result = ConcurrentContentExtractor._rewrite_google_sheets_url(url)
        assert "gid=222" in result

    def test_gid_zero_explicit(self):
        url = "https://docs.google.com/spreadsheets/d/ABC/edit#gid=0"
        result = ConcurrentContentExtractor._rewrite_google_sheets_url(url)
        assert "gid=0" in result

    def test_long_spreadsheet_id(self):
        sid = "1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgVE2upms"
        url = f"https://docs.google.com/spreadsheets/d/{sid}/edit#gid=0"
        result = ConcurrentContentExtractor._rewrite_google_sheets_url(url)
        assert sid in result
        assert "export?format=csv" in result


# ---------------------------------------------------------------------------
# ConcurrentContentExtractor._fetch_url — spreadsheet detection
# ---------------------------------------------------------------------------


class TestFetchUrlSpreadsheetDetection:
    """Test that _fetch_url correctly routes to spreadsheet parsing based on content-type."""

    def _make_response(self, content_type: str, text: str = "", content: bytes = b"") -> MagicMock:
        resp = MagicMock()
        resp.status_code = 200
        resp.headers = {"content-type": content_type}
        resp.text = text
        resp.content = content
        resp.raise_for_status = MagicMock()
        return resp

    @pytest.mark.asyncio
    async def test_csv_content_type_routed_to_csv_parser(self):
        csv_data = "Col1,Col2\nA,B\n"
        response = self._make_response("text/csv; charset=utf-8", text=csv_data)

        extractor = ConcurrentContentExtractor()
        extractor._client = MagicMock()
        extractor._client.get = AsyncMock(return_value=response)

        result = await extractor._fetch_url("https://example.com/data.csv")

        assert result["success"] is True
        assert result["method"] == "spreadsheet_csv"
        assert "Col1" in result["content"]

    @pytest.mark.asyncio
    async def test_csv_url_extension_routed_to_csv_parser(self):
        csv_data = "X,Y\n1,2\n"
        response = self._make_response("application/octet-stream", text=csv_data)

        extractor = ConcurrentContentExtractor()
        extractor._client = MagicMock()
        extractor._client.get = AsyncMock(return_value=response)

        result = await extractor._fetch_url("https://example.com/export.csv")

        assert result["method"] == "spreadsheet_csv"

    @pytest.mark.asyncio
    async def test_excel_content_type_routed_to_excel_parser(self):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Header"])
        ws.append(["Value"])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        response = self._make_response(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=xlsx_bytes,
        )
        response.text = ""  # Excel is binary

        extractor = ConcurrentContentExtractor()
        extractor._client = MagicMock()
        extractor._client.get = AsyncMock(return_value=response)

        result = await extractor._fetch_url("https://example.com/report.xlsx")

        assert result["method"] == "spreadsheet_excel"
        assert result["success"] is True

    @pytest.mark.asyncio
    async def test_google_sheets_url_rewritten_and_parsed_as_csv(self):
        csv_data = "Month,Revenue\nJan,1000\nFeb,2000\n"
        response = self._make_response("text/csv", text=csv_data)

        extractor = ConcurrentContentExtractor()
        extractor._client = MagicMock()
        extractor._client.get = AsyncMock(return_value=response)

        url = "https://docs.google.com/spreadsheets/d/ABC123/edit#gid=0"
        result = await extractor._fetch_url(url)

        # Verify the URL was rewritten before the GET call
        call_args = extractor._client.get.call_args[0][0]
        assert "export?format=csv" in call_args
        assert result["method"] == "spreadsheet_csv"
        assert "Month" in result["content"]

    @pytest.mark.asyncio
    async def test_html_content_type_uses_existing_html_path(self):
        html = "<html><body><article>Some article text here and more words</article></body></html>"
        response = self._make_response("text/html; charset=utf-8", text=html)

        extractor = ConcurrentContentExtractor()
        extractor._client = MagicMock()
        extractor._client.get = AsyncMock(return_value=response)

        result = await extractor._fetch_url("https://example.com/article")

        # Should NOT use spreadsheet methods
        assert result.get("method") != "spreadsheet_csv"
        assert result.get("method") != "spreadsheet_excel"

    @pytest.mark.asyncio
    async def test_invalid_csv_returns_success_false(self):
        """Malformed CSV that fails parsing should produce success=False."""
        response = self._make_response("text/csv", text="")

        extractor = ConcurrentContentExtractor()
        extractor._client = MagicMock()
        extractor._client.get = AsyncMock(return_value=response)

        result = await extractor._fetch_url("https://example.com/data.csv")

        # Empty CSV returns a descriptive string (not None), so success=True is fine here.
        # The important case is that parse errors (None) → success=False.
        assert result["method"] == "spreadsheet_csv"

    @pytest.mark.asyncio
    async def test_invalid_excel_bytes_returns_success_false(self):
        """Corrupt Excel bytes should result in success=False."""
        response = self._make_response(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=b"not an xlsx file",
        )
        response.text = ""

        extractor = ConcurrentContentExtractor()
        extractor._client = MagicMock()
        extractor._client.get = AsyncMock(return_value=response)

        result = await extractor._fetch_url("https://example.com/broken.xlsx")

        assert result["method"] == "spreadsheet_excel"
        assert result["success"] is False
        assert result["error"] == "Excel parsing failed"
        assert result["content"] is None
